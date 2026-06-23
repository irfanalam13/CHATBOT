"""Text extraction for every supported document type.

Returns (text, page_map) where page_map is [(char_offset, page_number)] used
to attribute chunks to source pages for citations.
"""
from __future__ import annotations

import csv
import io
import json
import xml.etree.ElementTree as ET

from app.core.exceptions import ValidationAppError


def extract(filename: str, data: bytes) -> tuple[str, list[tuple[int, int]]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    handler = {
        "pdf": _pdf,
        "docx": _docx,
        "txt": _plain,
        "md": _plain,
        "csv": _csv,
        "xlsx": _xlsx,
        "xls": _xlsx,
        "html": _html,
        "htm": _html,
        "json": _json,
        "xml": _xml,
        "eml": _email,
    }.get(ext)
    if not handler:
        raise ValidationAppError(f"Unsupported file type: .{ext}")
    return handler(data)


def _plain(data: bytes) -> tuple[str, list]:
    return data.decode("utf-8", errors="ignore"), []


def _pdf(data: bytes) -> tuple[str, list[tuple[int, int]]]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    page_map: list[tuple[int, int]] = []
    offset = 0
    for page_no, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        page_map.append((offset, page_no))
        parts.append(text)
        offset += len(text) + 1
    return "\n".join(parts), page_map


def _docx(data: bytes) -> tuple[str, list]:
    from docx import Document as Docx

    doc = Docx(io.BytesIO(data))
    text = "\n".join(p.text for p in doc.paragraphs)
    for table in doc.tables:
        for row in table.rows:
            text += "\n" + " | ".join(c.text for c in row.cells)
    return text, []


def _csv(data: bytes) -> tuple[str, list]:
    reader = csv.reader(io.StringIO(data.decode("utf-8", errors="ignore")))
    return "\n".join(" | ".join(row) for row in reader), []


def _xlsx(data: bytes) -> tuple[str, list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            lines.append(" | ".join("" if c is None else str(c) for c in row))
    return "\n".join(lines), []


def _html(data: bytes) -> tuple[str, list]:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(data, "lxml")
    for tag in soup(["script", "style"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True), []


def _json(data: bytes) -> tuple[str, list]:
    obj = json.loads(data.decode("utf-8", errors="ignore"))
    return json.dumps(obj, indent=2, ensure_ascii=False), []


def _xml(data: bytes) -> tuple[str, list]:
    root = ET.fromstring(data.decode("utf-8", errors="ignore"))
    parts: list[str] = []

    def walk(node: ET.Element) -> None:
        if node.text and node.text.strip():
            parts.append(f"{node.tag}: {node.text.strip()}")
        for child in node:
            walk(child)

    walk(root)
    return "\n".join(parts), []


def _email(data: bytes) -> tuple[str, list]:
    import mailparser

    mail = mailparser.parse_from_bytes(data)
    header = f"From: {mail.from_}\nTo: {mail.to}\nSubject: {mail.subject}\n\n"
    return header + (mail.body or ""), []
